import datetime as dt
import dateutil
import time
import glob
import os
from pathlib import Path
import re

import numpy as np
import pandas as pd
from pytz import timezone
from openpyxl import load_workbook
from sklearn.impute import KNNImputer
import statsmodels.api as sm
# from statsmodels.imputation.mice import MICE
import tqdm

import data
from constants import SEOUL_STATIONS

from constants import SEOUL_STATIONS, SEOULTZ

def stats_preprocess():
    print("Data preprocessing(imputation) start...")

    raw_df = data.load(datecol=[1])
    dfs_h = []

    impute_statistics = {}
    stat_dir = Path("/mnt/data/impute_stats")
    Path.mkdir(stat_dir, parents=True, exist_ok=True)
    for station_name in tqdm.tqdm(SEOUL_STATIONS.keys(), total=len(SEOUL_STATIONS.keys())):
        sdf = data.load_station(raw_df, SEOUL_STATIONS[station_name])

        df_isna = sdf.isna().sum()
        df_isnotna = sdf.notna().sum()
        df_isna.to_csv(stat_dir / f"stats_{station_name}_isna.csv")
        df_isnotna.to_csv(stat_dir / f"stats_{station_name}_isnotna.csv")

        imputer = KNNImputer(
            n_neighbors=5, weights="distance", missing_values=np.NaN)
        _df = pd.DataFrame(imputer.fit_transform(sdf))
        _df.columns = sdf.columns
        _df.index = sdf.index

        dfs_h.append(_df)

    df = pd.concat(dfs_h)
    df.to_csv("/input/python/input_seoul_imputed_hourly_pandas.csv")


def parse_obsxlsx(obs_path, input_dir):
    """Parse observatory xlsx file

    row[1] = state (시, 도)
    row[2] = city (도시)
    row[3] = stationCode (측정소코드)
    row[4] = stationName (측정소명)
    row[5] = address (주소)
    row[6] = langitude (경도)
    row[7] = latitude (위도)
    row[8] = remakrs (비고)
    """
    wb = load_workbook(str(obs_path))
    ws = wb.get_sheet_by_name(name = '2017년')

    re_opened = r"신규"
    re_closed = r"폐쇄"
    re_moved = r"이전"
    re_hangul = r"[가-힣]+"
    re_opened_date = r"(\d+)\.\W*(\d+)\.?"
    re_closed_date = r"(\d+)\.\W*(\d+)\.?(\W*(\d+)\.?)*"

    df = pd.DataFrame(columns=['stationCode', 'stationName',
        'oDate', 'cDate', 'lon', 'lat'])

    for row in ws.iter_rows(min_row=ws.min_row, max_row=ws.max_row,
        min_col=1, max_col=8, values_only=True):
        # just pick old date
        opened_date = dt.datetime(1970, 1, 1, 1).astimezone(SEOULTZ)

        # close value to maximum value in Int64
        closed_date = dt.datetime(2037, 12, 31, 23).astimezone(SEOULTZ)

        # if stationCode is blank, skip, so no move
        if row[3] == '측정소명' or row[2] == None or np.isnan(float(row[2])):
            continue

        # if there is not stationCode (i.e. wrong column), skip loop
        if isinstance(row[2], int):
            station_code = int(row[2])
        else:
            raise ValueError()

        lon_X = float(row[5])
        lat_Y = float(row[6])

        station_name = re.findall(re_hangul, row[3])[0]

        if re.match(re_opened, str(row[7])):
            opened = re.findall(re_opened_date, row[7])
            opened_year = int(opened[0][0])
            opened_month = int(opened[0][1])

            opened_date = dt.datetime(opened_year, opened_month, 2, 0)

        if re.match(re_closed, str(row[7])):
            closed = re.findall(re_closed_date, row[7])
            closed_year = int(closed[0][0])
            closed_month = int(closed[0][1])

            closed_date = dt.datetime(closed_year, closed_month, 1, 23)

        assert isinstance(station_code, int)
        assert isinstance(lon_X, float)
        assert isinstance(lat_Y, float)
        df = df.append(pd.DataFrame([[station_code, station_name, opened_date, closed_date, lon_X, lat_Y]], \
                columns=df.columns))

    df.set_index('stationCode', inplace=True)
    df.to_csv("obs.csv")
    return df

def parse_weathers(wea_dir, seoul_stn_code=108):
    re_wea_fn = r"SURFACE_ASOS_$(string(wea_stn_code))_HR_([0-9]+)_([0-9]+)_([0-9]+).csv"
    date_str = "yyyy-mm-dd HH:MM"

    # for tqdm, conver to list
    wea_globs = list(wea_dir.rglob('*.csv'))

    dfs = []
    imputer = KNNImputer(
            n_neighbors=5, weights="distance", missing_values=np.NaN)
    for wea_path in tqdm.tqdm(wea_globs):
        _df_raw = pd.read_csv(wea_path)
        _df_raw.rename(mapper={"일시": "date",
                    "기온(°C)": "temp",
                    "강수량(mm)": "prep",
                    "풍속(m/s)": "wind_spd",
                    "풍향(16방위)": "wind_dir",
                    "습도(%)": "humid",
                    "현지기압(hPa)": "pres",
                    "적설(cm)": "snow"}, axis='columns', inplace=True)

        _dates = [dateutil.parser.isoparse(str(d)).astimezone(SEOULTZ) for d in _df_raw.loc[:, 'date']]

        dict_wind_dir = {"0": 90,
            "360": 90 , "20" : 67.5 , "50" : 45 , "70" : 22.5 ,
            "90" : 0  , "110": 337.5, "140": 315, "160": 292.5,
            "180": 270, "200": 247.5, "230": 225, "250": 202.5,
            "270": 180, "290": 157.5, "320": 135, "340": 112.5}

        _df_raw['prep'] = _df_raw['prep'].replace(np.nan, 0.0)
        _df_raw['snow'] = _df_raw['snow'].replace(np.nan, 0.0)

        _df = pd.DataFrame({
                'date' : _dates,
                'temp' : _df_raw.loc[:, 'temp'],
                'wind_spd' : _df_raw.loc[:, 'wind_spd'], 'wind_dir' : _df_raw.loc[:, 'wind_dir'],
                # 'u' : _u, 'v' : _v,
                # 'wind_sdir' : _wind_sdir, 'wind_cdir' : _wind_cdir,
                'pres' : _df_raw.loc[:, 'pres'],
                'humid' : _df_raw.loc[:, 'humid'],
                'prep' : _df_raw.loc[:, 'prep'],
                'snow' : _df_raw.loc[:, 'snow']})

        _df.sort_values(by=['date'], inplace=True)
        _df.set_index('date', inplace=True)
        # imputation
        _df = pd.DataFrame(imputer.fit_transform(_df.values),
                           index=_df.index, columns=_df.columns)

        # if wind direction is imputed,
        # dict key may not match -> find nearest key
        def approx_winddir(w):
            wd = np.array([0, 20, 50, 70, 90, 110, 140, 160, 180, 200, 230, 250, 270, 290, 320, 340, 360])
            return wd[np.argmin(np.sqrt((w - wd)**2))]

        _wind_dir = np.array([approx_winddir(w) for w in _df.loc[:, 'wind_dir']])
        _df.loc[:, 'wind_dir'] = _wind_dir

        # Computation with imputed values
        _wind_cdir = [np.cos(np.deg2rad(270.0 - dict_wind_dir[str(int(w))])) \
                for w in _df.loc[:, 'wind_dir']]
        _wind_sdir = [np.sin(np.deg2rad(270.0 - dict_wind_dir[str(int(w))])) \
                for w in _df.loc[:, 'wind_dir']]

        _u = [w[0] * np.cos(np.deg2rad(270.0 - dict_wind_dir[str(int(w[1]))])) \
                for w in zip(_df.loc[:, 'wind_spd'], _df.loc[:, 'wind_dir'])]
        _v = [w[0] * np.sin(np.deg2rad(270.0 - dict_wind_dir[str(int(w[1]))])) \
                for w in zip(_df.loc[:, 'wind_spd'], _df.loc[:, 'wind_dir'])]
        # add column with computed column (wind)
        _df['u'] = _u
        _df['v'] = _v
        _df['wind_sdir'] = _wind_sdir
        _df['wind_cdir'] = _wind_cdir
        dfs.append(_df)

    df = pd.concat(dfs)
    df.sort_values(by=['date'], inplace=True)

    df.to_csv("wea.csv")
    return df

def parse_aerosols(aes_dir):
    re_aes_fn = r"**/([0-9]+)년\W*([0-9]+)(분기|월).csv"
    re_ext_fn = r".*.csv"
    date_str = "%Y%m%d%H"
    re_date = "^([0-9]{4})([0-9]{2})([0-9]{2})([0-9]{2})$"

    # for tqdm, conver to list
    aes_globs = list(aes_dir.rglob('*.csv'))

    dfs = []
    imputer = KNNImputer(
            n_neighbors=5, weights="distance", missing_values=np.NaN)

    for aes_path in tqdm.tqdm(aes_globs):
        _df_raw = pd.read_csv(aes_path)
        _df_raw.rename(mapper={"지역": "region", "측정소코드": "stationCode",
                    "측정소명": "stationName", "측정일시": "date",
                    "주소": "addr"}, axis='columns', inplace=True)

        def _parse_date(dstr, date_str):
            m = re.match(re_date, str(dstr))
            yyyy = int(m.groups()[0])
            mm = int(m.groups()[1])
            dd = int(m.groups()[2])
            hh = int(m.groups()[3])

            if hh == 24:
                d = dt.datetime(yyyy, mm, dd, hh-1).astimezone(SEOULTZ) + \
                    dt.timedelta(hours=1)
            else:
                d = dt.datetime(yyyy, mm, dd, hh).astimezone(SEOULTZ)

            return d

        _dates = list(map(lambda d: _parse_date(str(d), date_str), _df_raw.loc[:, 'date'].tolist()))

        # Imputation on whole aerosol dataset takes too long time,
        # DataFrame for imputation
        # _df_imp = pd.DataFrame({
        #         'stationCode' : _df_raw.loc[:, 'stationCode'],
        #         'date' : _dates,
        #         'SO2' : _df_raw.loc[:, 'SO2'],
        #         'CO' : _df_raw.loc[:, 'CO'],
        #         'O3' : _df_raw.loc[:, 'O3'],
        #         'NO2' : _df_raw.loc[:, 'NO2'],
        #         'PM10' : _df_raw.loc[:, 'PM10'],
        #         'PM25' : _df_raw.loc[:, 'PM25']})

        # _df_imp.sort_values(by=['stationCode', 'date'], inplace=True)
        # _df_imp.set_index(['stationCode', 'date'], inplace=True)
        # _df_imp = pd.DataFrame(imputer.fit_transform(_df_imp.values),
        #                    index=_df_imp.index, columns=_df_imp.columns)

        # imputed results to DataFrame
        _df = pd.DataFrame({
                'stationCode' : _df_raw.loc[:, 'stationCode'],
                'date' : _dates,
                'SO2' : _df_raw.loc[:, 'SO2'],
                'CO' : _df_raw.loc[:, 'CO'],
                'O3' : _df_raw.loc[:, 'O3'],
                'NO2' : _df_raw.loc[:, 'NO2'],
                'PM10' : _df_raw.loc[:, 'PM10'],
                'PM25' : _df_raw.loc[:, 'PM25']})

        _df.sort_values(by=['stationCode', 'date'], inplace=True)

        dfs.append(_df)

    df = pd.concat(dfs)
    df.sort_values(by=['date', 'stationCode'], inplace=True)
    df.set_index(['stationCode', 'date'], inplace=True)

    df.to_csv("aes.csv")
    return df

def stats_parse():
    seoul_stn_code = 108
    input_dir = Path("/input")
    obs_path = input_dir / "station" / "aerosol_observatory_2017_aironly.xlsx"
    aes_dir = input_dir / "aerosol"
    wea_dir = input_dir / "weather" / "seoul"

    print("Parsing Station dataset...", flush=True)
    df_obs = parse_obsxlsx(obs_path, input_dir)
    print("Parsing Weather dataset...", flush=True)
    df_wea = parse_weathers(wea_dir, seoul_stn_code=seoul_stn_code)
    print("Parsing Aerosol dataset...", flush=True)
    df_aes = parse_aerosols(aes_dir)

    # join three DataFrame
    df1 = df_aes.join(df_obs, on='stationCode')
    df2 = df1.join(df_wea, on='date', how='inner')

    df = df2.loc[:, ["lat", "lon",
                   "SO2", "CO", "O3", "NO2", "PM10", "PM25",
                   "temp", "pres",
                   "u", "v", "wind_spd", "wind_dir", "wind_sdir", "wind_cdir",
                   "humid", "prep", "snow"]]

    Path.mkdir(input_dir, parents=True, exist_ok=True)
    df.to_csv(input_dir / "input_2021.csv")

    # print("Finished Creating input.csv...", flush=True)
    # impute_seoul(df, input_dir)

def impute_seoul(df, input_dir):
    dfs_h = []
    for station_name in tqdm.tqdm(SEOUL_STATIONS.keys(), total=len(SEOUL_STATIONS.keys())):
        print(station_name)
        sdf = df.query('stationCode == "' + SEOUL_STATIONS[station_name] + '"')

        imputer = KNNImputer(
            n_neighbors=5, weights="distance", missing_values=np.NaN)
        _df = pd.DataFrame(imputer.fit_transform(sdf))
        _df.columns = sdf.columns
        _df.index = sdf.index

        dfs_h.append(_df)

    df = pd.concat(dfs_h)
    df.to_csv("/input/python/input_seoul_imputed_hourly_pandas.csv")
